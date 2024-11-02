import os
import pickle
import openpyxl
from openpyxl.styles import PatternFill
from concurrent.futures import ProcessPoolExecutor
from src.QuadTree import *


if __name__ == "__main__":
    # Read quadtree data from a pickle file
    with open("data/quadtrees.pkl", "rb") as f:
        quadtree_data = pickle.load(f)

    # Use ProcessPoolExecutor for parallel processing
    with ProcessPoolExecutor() as executor:
        pixel_data_of_video = list(executor.map(processQuadtree, quadtree_data[:]))

    # Path to the Excel file
    current_dir = os.getcwd()
    file_path = os.path.join(current_dir, "data", "MatrixView.xlsx")

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Create sheets for each frame
    for frame_index, pixel_data_of_image in enumerate(pixel_data_of_video):
        sheet = workbook.create_sheet(title=f"Frame {frame_index + 1}")
        for col in range(1, 3840):  # Assuming a maximum of 3840 columns
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 0.7
        for row in range(1, len(pixel_data_of_image) + 1):
            sheet.row_dimensions[row].height = 7

        # Create a dictionary to store color data for each cell
        fill_dict = {}
        for pixel_data in pixel_data_of_image:
            (x, y), color = pixel_data
            if isinstance(color, int):  # Convert to hex if color is an integer
                color = rgbToHex(color)
            excel_cell = (x + int(180 * 2.5), y + 90)
            fill_dict[excel_cell] = PatternFill(
                start_color=color[1:], end_color=color[1:], fill_type="solid"
            )

        # Apply color to each cell in the sheet in one pass
        for (x, y), fill in fill_dict.items():
            cell = sheet.cell(row=y, column=x)
            cell.fill = fill

    # Remove the default sheet
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # Save the workbook
    workbook.save(file_path)
    print(f"Excel file saved at: {file_path}")
